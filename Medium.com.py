from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

def fetch_titles_and_links(url, limit=100):
    titles_and_links = []

    # Configure Chrome options
    options = Options()
    options.add_argument("--headless")  # Run Chrome in headless mode (no GUI)
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("--no-sandbox")

    # Path to Chrome WebDriver executable
    webdriver_path = r'C:\Users\Himu\Desktop\New folder (3)\chromedriver-win64\chromedriver.exe'

    # Initialize Chrome WebDriver
    service = Service(webdriver_path)
    driver = webdriver.Chrome(service=service, options=options)

    try:
        # Load the URL
        driver.get(url)

        # Wait for the page to load all content
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "h2")))
        time.sleep(2)  # Add a short delay to ensure all content is loaded (adjust as needed)

        # Scroll down the page to load more content if necessary
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)  # Adjust delay as needed after scrolling

        # Parse HTML with BeautifulSoup
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Extract titles and links
        while len(titles_and_links) < limit:
            for h2_tag in soup.find_all('h2'):
                title = h2_tag.text.strip()

                # Find the link associated with the title
                link = h2_tag.find_next('a', href=True)
                if link:
                    href = link['href']
                    full_link = 'https://medium.com' + href
                else:
                    full_link = "Link not found"

                titles_and_links.append((title, full_link))

                # Break the loop if we reach the limit
                if len(titles_and_links) >= limit:
                    break

            # If we haven't reached the limit, try to scroll down to load more content
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(3)  # Adjust delay after scrolling

                # Update BeautifulSoup object with new content
                soup = BeautifulSoup(driver.page_source, 'html.parser')
            except Exception as e:
                print(f"Error scrolling or extracting data: {e}")
                break

    finally:
        # Quit the WebDriver session
        driver.quit()

    return titles_and_links[:limit]  # Ensure only up to `limit` entries are returned

def write_to_excel(titles_and_links):
    # Create a new Workbook
    wb = Workbook()

    # Activate the active worksheet
    ws = wb.active
    ws.title = "Medium Articles"  # Set the worksheet title

    # Write headers
    ws.cell(row=1, column=1, value="Title")
    ws.cell(row=1, column=2, value="Link")

    # Write data
    for idx, (title, link) in enumerate(titles_and_links, start=2):
        ws.cell(row=idx, column=1, value=title)
        ws.cell(row=idx, column=2, value=link)

    # Save the workbook
    wb.save("medium_articles.xlsx")
    print("Excel file 'medium_articles.xlsx' has been created successfully.")

if __name__ == "__main__":
    url = 'https://medium.com/tag/aws/archive'
    titles_and_links = fetch_titles_and_links(url, limit=100)

    # Write titles and links to Excel
    write_to_excel(titles_and_links)
